from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel,
    QTableWidget, QTableWidgetItem, QPushButton,
    QDateEdit, QMessageBox
)
from PySide6.QtCore import Qt, QDate
from openpyxl import Workbook
from openpyxl.styles import Font
from pathlib import Path
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
)
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors

from models.session_repo import (
    get_student_history,
    get_student_history_range
)
from utils.time_utils import format_duration
from utils.resource_utils import data_path


class StudentHistoryWindow(QDialog):
    def __init__(self, student_id: str, student_name: str):
        super().__init__()

        self.student_id = student_id
        self.student_name = student_name
        self.current_rows = []

        self.setWindowTitle("Student Attendance History")
        self.resize(820, 480)

        layout = QVBoxLayout(self)

        # ---------- TITLE ----------
        title = QLabel(f"ID: {student_id}    Name: {student_name}")
        title.setStyleSheet("font-size: 15px; font-weight: bold;")
        layout.addWidget(title)

        # ---------- FILTER BAR ----------
        filter_bar = QHBoxLayout()

        filter_bar.addWidget(QLabel("From:"))
        self.from_date = QDateEdit()
        self.from_date.setCalendarPopup(True)
        self.from_date.setDate(QDate.currentDate().addMonths(-1))
        filter_bar.addWidget(self.from_date)

        filter_bar.addWidget(QLabel("To:"))
        self.to_date = QDateEdit()
        self.to_date.setCalendarPopup(True)
        self.to_date.setDate(QDate.currentDate())
        filter_bar.addWidget(self.to_date)

        btn_filter = QPushButton("Apply Filter")
        btn_filter.clicked.connect(self.apply_filter)
        filter_bar.addWidget(btn_filter)

        filter_bar.addStretch()
        layout.addLayout(filter_bar)

        # ---------- TOTALS ----------
        self.lbl_totals = QLabel()
        self.lbl_totals.setStyleSheet("font-weight: bold;")
        layout.addWidget(self.lbl_totals)

        # ---------- TABLE ----------
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(
            ["Date", "Entry Time", "Exit Time", "Duration"]
        )
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setSelectionMode(QTableWidget.NoSelection)
        self.table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.table)

        # ---------- BUTTON BAR ----------
        btn_bar = QHBoxLayout()

        self.btn_export_xlsx = QPushButton("Export Excel")
        self.btn_export_xlsx.clicked.connect(self.export_view)
        btn_bar.addWidget(self.btn_export_xlsx)

        self.btn_export_pdf = QPushButton("Export PDF")
        self.btn_export_pdf.clicked.connect(self.export_pdf)
        btn_bar.addWidget(self.btn_export_pdf)

        btn_bar.addStretch()

        btn_close = QPushButton("Close")
        btn_close.clicked.connect(self.close)
        btn_bar.addWidget(btn_close)

        layout.addLayout(btn_bar)

        # Initial load
        self.load_rows(get_student_history(student_id))

    # ---------- LOAD ----------
    def load_rows(self, rows):
        self.current_rows = rows
        self.table.setRowCount(len(rows))

        total_seconds = 0

        for r, (d, start, end, dur) in enumerate(rows):
            self.table.setItem(r, 0, QTableWidgetItem(d or "-"))
            self.table.setItem(r, 1, QTableWidgetItem(start or "-"))
            self.table.setItem(r, 2, QTableWidgetItem(end or "-"))
            self.table.setItem(
                r, 3,
                QTableWidgetItem(format_duration(dur or 0))
            )
            total_seconds += dur or 0

        self.lbl_totals.setText(
            f"Sessions: {len(rows)}    "
            f"Total Time: {format_duration(total_seconds)}"
        )

    # ---------- FILTER ----------
    def apply_filter(self):
        start = self.from_date.date().toString("yyyy-MM-dd")
        end = self.to_date.date().toString("yyyy-MM-dd")

        rows = get_student_history_range(
            self.student_id,
            start,
            end
        )

        self.load_rows(rows)

    # ---------- EXPORT ----------
    def export_view(self):
        if not self.current_rows:
            QMessageBox.warning(
                self,
                "Export",
                "No data to export for the selected range."
            )
            return

        report_dir = Path(data_path("reports/student"))
        report_dir.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance History"

        bold = Font(bold=True)

        # ---- SUMMARY ----
        total_seconds = sum((d or 0) for _, _, _, d in self.current_rows)

        ws["A1"] = "Student ID"
        ws["B1"] = self.student_id
        ws["A1"].font = bold

        ws["A2"] = "Student Name"
        ws["B2"] = self.student_name
        ws["A2"].font = bold

        ws["A3"] = "Sessions"
        ws["B3"] = len(self.current_rows)
        ws["A3"].font = bold

        ws["A4"] = "Total Time"
        ws["B4"] = format_duration(total_seconds)
        ws["A4"].font = bold

        ws.append([])

        # ---- TABLE ----
        ws.append(["Date", "Entry Time", "Exit Time", "Duration"])
        ws.row_dimensions[6].font = bold

        for d, start, end, dur in self.current_rows:
            ws.append([
                d,
                start or "-",
                end or "-",
                format_duration(dur or 0)
            ])

        for col in ("A", "B", "C", "D"):
            ws.column_dimensions[col].width = 22

        file_path = report_dir / f"Student_{self.student_id}_Filtered.xlsx"
        wb.save(file_path)

        QMessageBox.information(
            self,
            "Export Successful",
            f"Filtered report exported:\n{file_path}"
        )

    def export_pdf(self):
        if not self.current_rows:
            QMessageBox.warning(
                self,
                "Export PDF",
                "No data to export for the selected range."
            )
            return

        report_dir = Path(data_path("reports/student"))
        report_dir.mkdir(parents=True, exist_ok=True)

        file_path = report_dir / f"Student_{self.student_id}_Filtered.pdf"

        doc = SimpleDocTemplate(
            str(file_path),
            pagesize=A4,
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36
        )

        styles = getSampleStyleSheet()
        story = []

        # ---------- TITLE ----------
        story.append(Paragraph(
            "<b>Student Attendance Report</b>",
            styles["Title"]
        ))
        story.append(Spacer(1, 12))

        # ---------- META ----------
        total_seconds = sum((d or 0) for _, _, _, d in self.current_rows)

        meta_text = f"""
        <b>Student ID:</b> {self.student_id}<br/>
        <b>Name:</b> {self.student_name}<br/>
        <b>From:</b> {self.from_date.date().toString("yyyy-MM-dd")}<br/>
        <b>To:</b> {self.to_date.date().toString("yyyy-MM-dd")}<br/>
        <b>Sessions:</b> {len(self.current_rows)}<br/>
        <b>Total Time:</b> {format_duration(total_seconds)}
        """

        story.append(Paragraph(meta_text, styles["Normal"]))
        story.append(Spacer(1, 16))

        # ---------- TABLE ----------
        table_data = [
            ["Date", "Entry Time", "Exit Time", "Duration"]
        ]

        for d, start, end, dur in self.current_rows:
            table_data.append([
                d,
                start or "-",
                end or "-",
                format_duration(dur or 0)
            ])

        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("FONT", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (1, 1), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))

        story.append(table)

        # ---------- BUILD ----------
        doc.build(story)

        QMessageBox.information(
            self,
            "Export PDF",
            f"PDF report exported successfully:\n{file_path}"
        )

