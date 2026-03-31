from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from models.database import get_connection
from utils.resource_utils import data_path


REPORT_DIR = Path(data_path("reports/daily"))


def export_daily_report(day: datetime.date) -> str:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)

    conn = get_connection()
    cur = conn.cursor()

    day_str = day.strftime("%Y-%m-%d")
    title_date = day.strftime("%d %b %Y")

    # ---------- SUMMARY ----------
    cur.execute("""
    WITH fixed AS (
        SELECT
            date(start_at, 'localtime') AS d,
            student_id,
            CASE
                WHEN end_at IS NOT NULL THEN duration_sec
                ELSE strftime('%s','now','localtime') - strftime('%s', start_at)
            END AS dur
        FROM sessions
        WHERE date(start_at) = date(?)
    )
    SELECT
        d,
        COUNT(*) AS total_users,
        ROUND(SUM(dur) / 3600.0, 2) AS hours_spent,
        COUNT(DISTINCT student_id) AS distinct_ids
    FROM fixed
    GROUP BY d;
    """, (day_str,))

    summary = cur.fetchone()

    # ---------- ATTENDANCE ----------
    cur.execute("""
    WITH fixed AS (
        SELECT
            s.student_id AS id,
            COALESCE(st.name, '') AS name,
            COALESCE(st.class, '') AS class,
            CASE
                WHEN s.end_at IS NOT NULL THEN s.duration_sec
                ELSE strftime('%s','now','localtime') - strftime('%s', s.start_at)
            END AS dur
        FROM sessions s
        LEFT JOIN students st ON st.student_id = s.student_id
        WHERE date(s.start_at) = date(?)
    )
    SELECT
        id,
        name,
        class,
        SUM(dur)
    FROM fixed
    GROUP BY id, name, class
    ORDER BY class, id;
    """, (day_str,))

    attendance = cur.fetchall()
    conn.close()

    if not attendance:
        raise ValueError("No attendance data found for the selected day")

    # ---------- EXCEL ----------
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Summary"

    ws1["A1"] = f"Library Daily Summary — {title_date}"
    ws1.merge_cells("A1:D1")
    ws1["A1"].font = Font(bold=True, size=16)
    ws1["A1"].alignment = Alignment(horizontal="center")

    ws1.append(["Date", "Total No. of Users", "No. of Hrs Spent", "Distinct IDs"])

    if summary:
        ws1.append([
            summary[0],
            summary[1],
            summary[2],
            summary[3]
        ])

    for col_letter in ("A", "B", "C", "D"):
        ws1.column_dimensions[col_letter].width = 22


    # ---------- SHEET 2 ----------
    ws2 = wb.create_sheet("Attendance")
    ws2["A1"] = f"Attendance — {title_date}"
    ws2.merge_cells("A1:D1")
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A1"].alignment = Alignment(horizontal="center")

    ws2.append(["ID", "Name", "Class", "Time Spent"])

    for sid, name, cls, dur in attendance:
        h = dur // 3600
        m = (dur % 3600) // 60
        ws2.append([sid, name, cls, f"{h}h {m}m"])

    for col_letter in ("A", "B", "C", "D"):
        ws2.column_dimensions[col_letter].width = 22


    file_path = REPORT_DIR / f"Daily_{day.strftime('%Y_%m_%d')}.xlsx"
    wb.save(file_path)

    return str(file_path)
