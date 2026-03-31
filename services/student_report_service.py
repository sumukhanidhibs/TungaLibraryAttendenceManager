from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font
from models.database import get_connection
from utils.time_utils import format_duration

REPORT_DIR = Path("reports/student")


def export_student_report(student_id: str) -> str:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)

    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            date(start_at),
            time(start_at),
            time(end_at),
            duration_sec
        FROM sessions
        WHERE student_id = ?
        ORDER BY start_at
    """, (student_id,))

    rows = cur.fetchall()
    conn.close()

    if not rows:
        raise ValueError("No attendance records found for this student")

    # ---------- TOTALS ----------
    total_visits = len(rows)
    total_seconds = sum((dur or 0) for _, _, _, dur in rows)

    first_date = rows[0][0]
    last_date = rows[-1][0]

    # ---------- EXCEL ----------
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance History"

    bold = Font(bold=True)

    # ---- SUMMARY BLOCK ----
    ws["A1"] = "Student ID"
    ws["B1"] = student_id
    ws["A1"].font = bold

    ws["A2"] = "Total Visits"
    ws["B2"] = total_visits
    ws["A2"].font = bold

    ws["A3"] = "Total Time Spent"
    ws["B3"] = format_duration(total_seconds)
    ws["A3"].font = bold

    ws["A4"] = "First Visit"
    ws["B4"] = first_date
    ws["A4"].font = bold

    ws["A5"] = "Last Visit"
    ws["B5"] = last_date
    ws["A5"].font = bold

    # ---- TABLE HEADER ----
    start_row = 7
    ws.append([])  # spacing
    ws.append(["Date", "Entry Time", "Exit Time", "Duration"])
    ws.row_dimensions[start_row].font = bold

    # ---- TABLE DATA ----
    for d, start, end, dur in rows:
        ws.append([
            d,
            start or "-",
            end or "-",
            format_duration(dur or 0)
        ])

    # ---- COLUMN WIDTHS ----
    for col in ("A", "B", "C", "D"):
        ws.column_dimensions[col].width = 22

    file_path = REPORT_DIR / f"Student_{student_id}.xlsx"
    wb.save(file_path)

    return str(file_path)


